from __future__ import annotations

import argparse
import csv
import json
from collections import deque
from functools import lru_cache
from pathlib import Path

import numpy as np
from PIL import Image, ImageDraw, ImageFilter, ImageFont
from openpyxl import Workbook, load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
EMBEDDED_PORTRAIT_DIR = PROJECT_ROOT / "outputs" / "embedded_portraits"
REQUIRED_COLUMNS = ("讲师姓名", "项目经历", "底部介绍", "人像照片")
OPTIONAL_COLUMNS = ("output_filename",)
COLUMN_ALIASES = {
    "讲师姓名": "name",
    "姓名": "name",
    "name": "name",
    "项目经历": "project_experience",
    "项目经理": "project_experience",
    "project_experience": "project_experience",
    "底部介绍": "bio",
    "讲师介绍": "bio",
    "介绍": "bio",
    "bio": "bio",
    "人像照片": "portrait_path",
    "人物照片": "portrait_path",
    "人像": "portrait_path",
    "portrait_path": "portrait_path",
    "输出文件名": "output_filename",
    "文件名": "output_filename",
    "output_filename": "output_filename",
}
REMBG_SESSION = None


def resolve_path(value: str) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    if value.startswith("../../assets/"):
        return (PROJECT_ROOT / value.replace("../../", "", 1)).resolve()
    return (PROJECT_ROOT / value).resolve()


@lru_cache(maxsize=8)
def load_config(template: str) -> dict:
    config_path = PROJECT_ROOT / "src" / "templates" / template / "config.json"
    with config_path.open("r", encoding="utf-8") as f:
        return json.load(f)


@lru_cache(maxsize=32)
def font(path: Path, size: int | float) -> ImageFont.FreeTypeFont:
    return ImageFont.truetype(str(path), size=round(size))


@lru_cache(maxsize=16)
def cached_rgba_image(path: Path) -> Image.Image:
    return Image.open(path).convert("RGBA")


def text_width(draw: ImageDraw.ImageDraw, text: str, ft: ImageFont.FreeTypeFont) -> int:
    if not text:
        return 0
    box = draw.textbbox((0, 0), text, font=ft)
    return box[2] - box[0]


def tracking_px(value: float, ft: ImageFont.FreeTypeFont) -> float:
    return float(value) * ft.size / 1000


def text_width_tracked(draw: ImageDraw.ImageDraw, text: str, ft: ImageFont.FreeTypeFont, tracking: float = 0) -> float:
    if not text:
        return 0
    width = sum(text_width(draw, ch, ft) for ch in text)
    return width + max(0, len(text) - 1) * tracking_px(tracking, ft)


def draw_tracked_text(
    draw: ImageDraw.ImageDraw,
    xy: tuple[float, float],
    text: str,
    ft: ImageFont.FreeTypeFont,
    fill,
    tracking: float = 0,
) -> None:
    x, y = xy
    gap = tracking_px(tracking, ft)
    for ch in text:
        draw.text((x, y), ch, font=ft, fill=fill)
        x += text_width(draw, ch, ft) + gap


def wrap_text(
    draw: ImageDraw.ImageDraw,
    text: str,
    ft: ImageFont.FreeTypeFont,
    max_width: int,
    max_lines: int,
    tracking: float = 0,
) -> list[str]:
    pieces = [part.strip() for part in text.splitlines() if part.strip()]

    lines: list[str] = []
    for piece in pieces:
        current = ""
        for ch in piece:
            candidate = current + ch
            if text_width_tracked(draw, candidate, ft, tracking) <= max_width:
                current = candidate
                continue
            if current:
                lines.append(current)
            current = ch
            if max_lines and len(lines) >= max_lines:
                return lines[:max_lines]
        if current:
            lines.append(current)
        if max_lines and len(lines) >= max_lines:
            return lines[:max_lines]
    return lines[:max_lines] if max_lines else lines


def manual_lines(text: str) -> list[str]:
    return [line.strip() for line in text.splitlines() if line.strip()]


def draw_tracked_text_clipped(
    base: Image.Image,
    clip_box: tuple[int, int, int, int],
    xy: tuple[float, float],
    text: str,
    ft: ImageFont.FreeTypeFont,
    fill,
    tracking: float = 0,
    shadow: dict | None = None,
) -> None:
    layer = Image.new("RGBA", base.size, (0, 0, 0, 0))
    layer_draw = ImageDraw.Draw(layer)
    if shadow:
        sx, sy = shadow.get("offset", [0, 0])
        opacity = float(shadow.get("opacity", 1))
        color = shadow.get("fill", "#000000")
        alpha = max(0, min(255, int(255 * opacity)))
        rgba = tuple(int(color.strip("#")[i : i + 2], 16) for i in (0, 2, 4)) + (alpha,)
        draw_tracked_text(layer_draw, (xy[0] + sx, xy[1] + sy), text, ft, rgba, tracking)
    draw_tracked_text(layer_draw, xy, text, ft, fill, tracking)

    mask = Image.new("L", base.size, 0)
    mask_draw = ImageDraw.Draw(mask)
    mask_draw.rectangle(clip_box, fill=255)
    base.alpha_composite(Image.composite(layer, Image.new("RGBA", base.size, (0, 0, 0, 0)), mask))


def draw_text_with_shadow(
    draw: ImageDraw.ImageDraw,
    xy: tuple[int, int],
    text: str,
    ft: ImageFont.FreeTypeFont,
    fill: str,
    shadow: dict | None = None,
    tracking: float = 0,
) -> None:
    if shadow:
        sx, sy = shadow.get("offset", [0, 0])
        opacity = float(shadow.get("opacity", 1))
        color = shadow.get("fill", "#000000")
        alpha = max(0, min(255, int(255 * opacity)))
        rgba = tuple(int(color.strip("#")[i : i + 2], 16) for i in (0, 2, 4)) + (alpha,)
        draw_tracked_text(draw, (xy[0] + sx, xy[1] + sy), text, ft, rgba, tracking)
    draw_tracked_text(draw, xy, text, ft, fill, tracking)


def texture_text(
    base: Image.Image,
    cfg: dict,
    text: str,
    ft: ImageFont.FreeTypeFont,
    texture_path: Path,
    tracking: float,
) -> None:
    mask = Image.new("L", base.size, 0)
    mask_draw = ImageDraw.Draw(mask)
    area = cfg["area"]
    left, top, right, bottom = area
    text_w = text_width_tracked(mask_draw, text, ft, tracking)
    bbox = mask_draw.textbbox((0, 0), text, font=ft)
    x = left + max(0, (right - left - text_w) / 2)
    if "bottom_y" in cfg:
        y = cfg["bottom_y"] - bbox[3]
    else:
        text_h = bbox[3] - bbox[1]
        y = top + max(0, (bottom - top - text_h) / 2) - bbox[1]
    draw_tracked_text(mask_draw, (x, y), text, ft, 255, tracking)

    text_bbox = mask.getbbox()
    if text_bbox is None:
        return

    texture = cached_rgba_image(texture_path)
    tiled = Image.new("RGBA", base.size, (0, 0, 0, 0))
    for tile_y in range(text_bbox[1], text_bbox[3], texture.height):
        for tile_x in range(text_bbox[0], text_bbox[2], texture.width):
            tiled.alpha_composite(texture, (tile_x, tile_y))
    clipped = Image.composite(tiled, Image.new("RGBA", base.size, (0, 0, 0, 0)), mask)
    base.alpha_composite(clipped)


def paste_portrait(canvas: Image.Image, portrait_path: Path, cfg: dict) -> None:
    portrait = prepare_portrait(Image.open(portrait_path))
    left, top, right, bottom = cfg["box"]
    target_w = right - left
    target_h = bottom - top
    if cfg.get("fit") == "contain_subject":
        fitted = fit_subject_to_box(portrait, target_w, target_h, cfg)
        paste_x = left + (target_w - fitted.width) // 2
        paste_y = portrait_top_y(fitted.height, bottom, cfg)
        paste_portrait_layer(canvas, fitted, (paste_x, paste_y), cfg)
        return

    scale = max(target_w / portrait.width, target_h / portrait.height)
    resized = portrait.resize((int(portrait.width * scale), int(portrait.height * scale)), Image.LANCZOS)
    crop_left = max(0, (resized.width - target_w) // 2)
    crop_top = max(0, resized.height - target_h)
    cropped = resized.crop((crop_left, crop_top, crop_left + target_w, crop_top + target_h))
    paste_portrait_layer(canvas, cropped, (left, top), cfg)


def paste_portrait_layer(canvas: Image.Image, portrait: Image.Image, xy: tuple[int, int], cfg: dict) -> None:
    layer = Image.new("RGBA", canvas.size, (0, 0, 0, 0))
    layer.alpha_composite(portrait, xy)
    layer = apply_portrait_layer_mask(layer, cfg)
    canvas.alpha_composite(layer)


def apply_portrait_layer_mask(layer: Image.Image, cfg: dict) -> Image.Image:
    mask_cfg = cfg.get("layer_mask")
    if not mask_cfg:
        return layer

    fade_start = int(mask_cfg.get("fade_start_y", 1032))
    fade_end = int(mask_cfg.get("fade_end_y", fade_start))
    if fade_end <= fade_start:
        return layer

    rgba = np.array(layer.convert("RGBA"))
    alpha = rgba[:, :, 3].astype(np.float32)
    height = alpha.shape[0]
    for y in range(max(0, fade_start), min(height, fade_end)):
        factor = 1 - ((y - fade_start) / (fade_end - fade_start))
        alpha[y, :] *= factor
    if fade_end < height:
        alpha[fade_end:, :] = 0
    rgba[:, :, 3] = np.clip(alpha, 0, 255).astype(np.uint8)
    return Image.fromarray(rgba, mode="RGBA")


def fit_subject_to_box(portrait: Image.Image, target_w: int, target_h: int, cfg: dict) -> Image.Image:
    alpha_threshold = int(cfg.get("alpha_threshold", 12))
    alpha = portrait.getchannel("A")
    mask = alpha.point(lambda value: 255 if value > alpha_threshold else 0)
    bbox = mask.getbbox() or (0, 0, portrait.width, portrait.height)
    subject = portrait.crop(bbox)
    subject = cleanup_alpha(subject, threshold=alpha_threshold)

    fill_ratio = float(cfg.get("fill_ratio", 0.98))
    max_h = target_h
    title_bottom = cfg.get("title_bottom_y")
    gap_cfg = cfg.get("head_gap")
    if title_bottom is not None and gap_cfg:
        max_h = min(max_h, cfg["box"][3] - (title_bottom + gap_cfg.get("min", 22)))
    scale = min(target_w / subject.width, max_h / subject.height) * fill_ratio
    new_size = (max(1, int(subject.width * scale)), max(1, int(subject.height * scale)))
    return subject.resize(new_size, Image.LANCZOS)


def portrait_top_y(subject_h: int, box_bottom: int, cfg: dict) -> int:
    bottom_anchor_y = box_bottom - subject_h
    title_bottom = cfg.get("title_bottom_y")
    gap_cfg = cfg.get("head_gap")
    if title_bottom is None or not gap_cfg:
        return bottom_anchor_y

    min_gap = gap_cfg.get("min", 22)
    max_gap = gap_cfg.get("max", 60)
    preferred = gap_cfg.get("preferred", (min_gap + max_gap) / 2)

    preferred_top = round(title_bottom + preferred)
    if preferred_top + subject_h <= box_bottom:
        return preferred_top

    top = box_bottom - subject_h
    gap = top - title_bottom
    if gap < min_gap:
        return round(title_bottom + min_gap)
    if gap > max_gap:
        return round(title_bottom + max_gap)
    return round(top)


def prepare_portrait(image: Image.Image) -> Image.Image:
    portrait = image.convert("RGBA")
    alpha = portrait.getchannel("A")
    if alpha.getextrema()[0] < 250:
        return portrait
    return remove_background(portrait)


def remove_background(image: Image.Image) -> Image.Image:
    try:
        return remove_background_rembg(image)
    except Exception:
        return remove_plain_background(image)


def remove_background_rembg(image: Image.Image) -> Image.Image:
    global REMBG_SESSION
    from rembg import new_session, remove

    if REMBG_SESSION is None:
        REMBG_SESSION = new_session("u2net_human_seg")
    result = remove(
        image.convert("RGB"),
        session=REMBG_SESSION,
        post_process_mask=True,
    ).convert("RGBA")
    return cleanup_alpha(result, threshold=18, erode=True, decontaminate=True)


def cleanup_alpha(image: Image.Image, threshold: int = 10, erode: bool = False, decontaminate: bool = False) -> Image.Image:
    rgba = np.array(image.convert("RGBA"))
    alpha = rgba[:, :, 3]
    alpha = np.where(alpha <= threshold, 0, alpha)
    alpha = np.where(alpha >= 248, 255, alpha)
    alpha_image = Image.fromarray(alpha.astype(np.uint8), mode="L")
    if erode:
        alpha_image = alpha_image.filter(ImageFilter.MinFilter(3))
    alpha_image = alpha_image.filter(ImageFilter.GaussianBlur(0.35))
    cleaned = image.convert("RGBA")
    if decontaminate:
        rgba = np.array(cleaned)
        alpha_f = np.array(alpha_image).astype(np.float32) / 255
        edge = (alpha_f > 0) & (alpha_f < 0.98)
        matte = np.array([18, 15, 10], dtype=np.float32)
        blend = np.power(alpha_f[..., None], 0.7)
        rgb = rgba[:, :, :3].astype(np.float32)
        rgb[edge] = rgb[edge] * blend[edge] + matte * (1 - blend[edge])
        rgba[:, :, :3] = np.clip(rgb, 0, 255).astype(np.uint8)
        cleaned = Image.fromarray(rgba, mode="RGBA")
    cleaned.putalpha(alpha_image)
    return cleaned


def remove_plain_background(image: Image.Image) -> Image.Image:
    rgba = np.array(image.convert("RGBA"))
    rgb = rgba[:, :, :3].astype(np.int16)
    height, width = rgb.shape[:2]
    edge = max(8, min(width, height) // 40)
    border = np.concatenate(
        [
            rgb[:edge, :, :].reshape(-1, 3),
            rgb[-edge:, :, :].reshape(-1, 3),
            rgb[:, :edge, :].reshape(-1, 3),
            rgb[:, -edge:, :].reshape(-1, 3),
        ],
        axis=0,
    )
    bg = np.median(border, axis=0)
    dist = np.linalg.norm(rgb - bg, axis=2)
    border_dist = np.linalg.norm(border - bg, axis=1)
    threshold = max(32, float(np.percentile(border_dist, 90) + 28))
    bg_candidate = dist <= threshold

    visited = np.zeros((height, width), dtype=bool)
    background = np.zeros((height, width), dtype=bool)
    queue: deque[tuple[int, int]] = deque()
    for x in range(width):
        for y in (0, height - 1):
            if bg_candidate[y, x] and not visited[y, x]:
                visited[y, x] = True
                queue.append((x, y))
    for y in range(height):
        for x in (0, width - 1):
            if bg_candidate[y, x] and not visited[y, x]:
                visited[y, x] = True
                queue.append((x, y))

    while queue:
        x, y = queue.popleft()
        background[y, x] = True
        for nx, ny in ((x + 1, y), (x - 1, y), (x, y + 1), (x, y - 1)):
            if 0 <= nx < width and 0 <= ny < height and not visited[ny, nx] and bg_candidate[ny, nx]:
                visited[ny, nx] = True
                queue.append((nx, ny))

    alpha = np.full((height, width), 255, dtype=np.uint8)
    alpha[background] = 0
    alpha_image = Image.fromarray(alpha, mode="L").filter(ImageFilter.GaussianBlur(1))
    image.putalpha(alpha_image)
    return image


def draw_project_experience(canvas: Image.Image, draw: ImageDraw.ImageDraw, cfg: dict, font_path: Path, text: str) -> None:
    title_cfg = cfg["title"]
    body_cfg = cfg["body"]
    title_font = font(font_path, title_cfg["font_size"])
    body_font = font(font_path, body_cfg["font_size"])

    left, top, right, bottom = body_cfg["box"]
    line_height = float(body_cfg["line_height"])
    tracking = body_cfg.get("tracking", 0)
    lines = manual_lines(text)
    ascent, _ = body_font.getmetrics()
    baseline = body_cfg["bottom_baseline"]
    first_y = baseline - (len(lines) - 1) * line_height - ascent

    title_y = first_y - line_height
    draw_tracked_text(
        draw,
        (title_cfg["xy"][0], title_y),
        title_cfg["text"],
        title_font,
        title_cfg["fill"],
        title_cfg.get("tracking", 0),
    )
    for idx, line in enumerate(lines):
        y = first_y + idx * line_height
        draw_tracked_text_clipped(
            canvas,
            (left, top, right, bottom),
            (left, int(y)),
            line,
            body_font,
            body_cfg["fill"],
            tracking,
            body_cfg.get("shadow"),
        )


def draw_bio_and_footer(
    canvas: Image.Image,
    draw: ImageDraw.ImageDraw,
    cfg: dict,
    font_path: Path,
    footer_font_path: Path,
    bio: str,
) -> None:
    bio_cfg = cfg["bio"]
    footer_cfg = cfg["footer"]
    bio_font = font(font_path, bio_cfg["font_size"])
    footer_font = font(footer_font_path, footer_cfg["font_size"])

    box_w = bio_cfg["width"]
    left = int((cfg["canvas"]["width"] - box_w) / 2)
    right = left + box_w
    pad_l, pad_t, pad_r, pad_b = bio_cfg["padding"]
    max_width = box_w - pad_l - pad_r
    bio_tracking = bio_cfg.get("tracking", 0)
    lines = wrap_text(draw, bio, bio_font, max_width, bio_cfg["max_lines"], bio_tracking)

    line_height = float(bio_cfg["line_height"])
    line_heights = [
        draw.textbbox((0, 0), line, font=bio_font)[3] - draw.textbbox((0, 0), line, font=bio_font)[1]
        for line in lines
    ]
    tallest_line = max(line_heights) if line_heights else 0
    text_block_h = (len(lines) - 1) * line_height + tallest_line if lines else 0
    box_h = int(round(text_block_h + pad_t + pad_b))

    footer_text = footer_cfg["text"]
    footer_bbox = draw.textbbox((0, 0), footer_text, font=footer_font)
    footer_h = footer_bbox[3] - footer_bbox[1]
    name_bottom = cfg["bottom_layout"]["name_group_bottom_y"]
    poster_bottom = cfg["bottom_layout"]["poster_bottom_y"]
    gap = (poster_bottom - name_bottom - box_h - footer_h) / 3

    box_top = name_bottom + gap
    box_bottom = box_top + box_h
    footer_y = box_bottom + gap - footer_bbox[1]

    draw.rounded_rectangle(
        (left, round(box_top), right, round(box_bottom)),
        radius=bio_cfg["radius"],
        outline=bio_cfg["stroke"],
        width=bio_cfg["stroke_width"],
    )
    draw_bio_glow_bars(canvas, cfg, left, right, box_top, box_bottom)

    text_y = box_top + (box_h - text_block_h) / 2
    for idx, line in enumerate(lines):
        x = left + pad_l
        y = text_y + idx * line_height
        draw_tracked_text(draw, (x, y), line, bio_font, bio_cfg["fill"], bio_tracking)

    footer_w = text_width_tracked(draw, footer_text, footer_font, footer_cfg.get("tracking", 0))
    footer_x = (cfg["canvas"]["width"] - footer_w) / 2
    draw_tracked_text(draw, (footer_x, footer_y), footer_text, footer_font, footer_cfg["fill"], footer_cfg.get("tracking", 0))


def draw_bio_glow_bars(canvas: Image.Image, cfg: dict, left: int, right: int, box_top: float, box_bottom: float) -> None:
    glow_cfg = cfg["bio"].get("glow_bar", {})
    if not glow_cfg.get("enabled", False):
        return
    glow_path = cfg["assets"].get("bio_box_glow_bar")
    if not glow_path:
        return
    glow = cached_rgba_image(resolve_path(glow_path)).copy()
    max_width = right - left
    if glow.width > max_width:
        scale = max_width / glow.width
        glow = glow.resize((max_width, max(1, int(glow.height * scale))), Image.LANCZOS)
    x = left + (max_width - glow.width) // 2
    offset_y = int(glow_cfg.get("offset_y", 0))
    top_y = round(box_top - glow.height / 2 + offset_y)
    bottom_y = round(box_bottom - glow.height / 2 + offset_y)
    canvas.alpha_composite(glow, (x, top_y))
    canvas.alpha_composite(glow, (x, bottom_y))


def render_teacher_image(config: dict, row: dict) -> Image.Image:
    expected_size = (config["canvas"]["width"], config["canvas"]["height"])
    base = cached_rgba_image(resolve_path(config["assets"]["base"]))
    if base.size != expected_size:
        raise ValueError(f"Base template size {base.size} does not match configured canvas {expected_size}.")

    canvas = base.copy()
    font_path = resolve_path(config["assets"]["font"])
    footer_font_path = resolve_path(config["assets"]["footer_font"])

    paste_portrait(canvas, resolve_path(row["portrait_path"]), config["portrait"])

    mask = cached_rgba_image(resolve_path(config["assets"]["foreground_mask"]))
    if mask.size != expected_size:
        raise ValueError(f"Foreground mask size {mask.size} does not match configured canvas {expected_size}.")
    canvas.alpha_composite(mask, (0, 0))

    draw = ImageDraw.Draw(canvas)
    draw_project_experience(canvas, draw, config["project_experience"], font_path, row["project_experience"])

    name_font = font(font_path, config["name"]["font_size"])
    tracking = config["name"]["tracking_by_length"].get(str(len(row["name"])), config["name"]["default_tracking"])
    texture_text(canvas, config["name"], row["name"], name_font, resolve_path(config["assets"]["name_texture"]), tracking)

    draw = ImageDraw.Draw(canvas)
    draw_bio_and_footer(canvas, draw, config, font_path, footer_font_path, row["bio"])

    return canvas


def save_rendered_image(image: Image.Image, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    suffix = output_path.suffix.lower()
    if suffix in {".jpg", ".jpeg"}:
        image.convert("RGB").save(output_path, quality=96)
        return
    image.save(output_path)


def render_teacher(config: dict, row: dict, output_dir: Path) -> Path:
    canvas = render_teacher_image(config, row)
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = clean_filename(row.get("output_filename") or f"{row['name']}_四星讲师.png")
    output_path = output_dir / filename
    save_rendered_image(canvas, output_path)
    return output_path


def clean_filename(filename: str) -> str:
    filename = str(filename).strip()
    for ch in '/\\:*?"<>|':
        filename = filename.replace(ch, "_")
    return filename


def read_csv_rows(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return [normalize_row(row) for row in csv.DictReader(f)]


def read_xlsx_rows(path: Path) -> list[dict]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    header = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    embedded_images = extract_xlsx_images(sheet, path)
    rows: list[dict] = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value is not None and str(value).strip() for value in values):
            continue
        row = {}
        for key, value in zip(header, values):
            if key:
                row[key] = "" if value is None else str(value)
        normalized = normalize_row(row)
        if not normalized.get("portrait_path") and row_number in embedded_images:
            normalized["portrait_path"] = str(embedded_images[row_number])
        rows.append(normalized)
    return rows


def extract_xlsx_images(sheet, source_path: Path) -> dict[int, Path]:
    images_by_row: dict[int, Path] = {}
    images = getattr(sheet, "_images", [])
    if not images:
        return images_by_row
    EMBEDDED_PORTRAIT_DIR.mkdir(parents=True, exist_ok=True)
    stem = source_path.stem
    for index, image in enumerate(images, start=1):
        anchor = getattr(image, "anchor", None)
        marker = getattr(anchor, "_from", None)
        if marker is None:
            continue
        row_number = marker.row + 1
        suffix = Path(getattr(image, "path", "")).suffix or ".png"
        output_path = EMBEDDED_PORTRAIT_DIR / f"{stem}_row{row_number}_{index}{suffix}"
        try:
            output_path.write_bytes(image._data())
        except Exception:
            continue
        images_by_row[row_number] = output_path
    return images_by_row


def normalize_row(row: dict) -> dict:
    normalized = {}
    for key, value in row.items():
        target_key = COLUMN_ALIASES.get(str(key).strip(), str(key).strip())
        normalized[target_key] = "" if value is None else str(value)
    return normalized


def read_rows(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_rows(path)
    raise ValueError(f"Unsupported input file type: {path.suffix}. Use .csv or .xlsx.")


def validate_rows(rows: list[dict]) -> None:
    if not rows:
        raise ValueError("Input file has no teacher rows.")
    missing_columns = [column for column in ("name", "project_experience", "bio", "portrait_path") if column not in rows[0]]
    if missing_columns:
        raise ValueError(f"Input file is missing required columns: {', '.join(missing_columns)}")
    for index, row in enumerate(rows, start=2):
        missing_values = [column for column in ("name", "project_experience", "bio", "portrait_path") if not str(row.get(column, "")).strip()]
        if missing_values:
            raise ValueError(f"Row {index} has empty required values: {', '.join(missing_values)}")


def with_output_format(filename: str, output_format: str) -> str:
    path = Path(filename)
    return f"{path.stem}.{output_format.lower()}"


def save_sample_xlsx(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "teachers"
    sheet.append(list(REQUIRED_COLUMNS))
    sheet.append(
        [
            "李甲",
            "服务10+主机厂和区域的新媒体提升辅导\n问界打深耕深圳站集训及多店全链路入店提升辅导\n比亚迪王朝云贵川战区新媒体直播获客专项辅导",
            "实战派辅导老师，14年汽车行业经验，6年新媒体经验，擅长新媒体全链路SOP集训/入店辅导，能快速切入门店痛点，用大白话讲专业方法，帮助学员拿到更有效的本地线索和成交闭环。",
            "assets/portraits/李甲_transparent.png",
        ]
    )
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="input/sample_teachers.csv")
    parser.add_argument("--template", default="four_star")
    parser.add_argument("--output-dir", default="outputs")
    parser.add_argument("--format", choices=["png", "jpg", "jpeg"], default=None)
    parser.add_argument("--make-sample-xlsx", action="store_true")
    args = parser.parse_args()

    if args.make_sample_xlsx:
        sample_path = PROJECT_ROOT / "input" / "teachers.xlsx"
        save_sample_xlsx(sample_path)
        print(sample_path)
        return

    config = load_config(args.template)
    rows = read_rows((PROJECT_ROOT / args.input).resolve())
    validate_rows(rows)
    for row in rows:
        if args.format:
            row["output_filename"] = with_output_format(row.get("output_filename") or f"{row['name']}_四星讲师.png", args.format)
        output_path = render_teacher(config, row, (PROJECT_ROOT / args.output_dir).resolve())
        print(output_path)


if __name__ == "__main__":
    main()
